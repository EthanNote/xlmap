import openpyxl
import json


def get_rect_from_range(range_str):
    data_range = range_str.split(':')
    left_top = openpyxl.utils.coordinate_to_tuple(data_range[0])
    right_bottom = openpyxl.utils.coordinate_to_tuple(data_range[1])
    return left_top[0]-1, left_top[1]-1, right_bottom[0], right_bottom[1]


def get_border_lines(border, rect):
    result = []

    left = 1
    top = 0
    right = 3
    bottom = 2

    if border.top.style:
        result.append((rect[top], rect[left], rect[top], rect[right]))

    if border.bottom.style:
        result.append((rect[bottom], rect[left], rect[bottom],  rect[right]))

    if border.left.style:
        result.append((rect[top], rect[left], rect[bottom],  rect[left]))

    if border.right.style:
        result.append((rect[top], rect[right], rect[bottom], rect[right]))

    if border.diagonalDown:
        result.append((rect[top], rect[left], rect[bottom], rect[right]))

    if border.diagonalUp:
        result.append((rect[bottom], rect[left], rect[top], rect[right]))

    return result


unit_width = 1
unit_height = 1


def get_real_height(lineHeight):
    return lineHeight/unit_height


def get_real_width(rowWidth):
    return rowWidth/unit_width


def get_xl_height(lineHeight):
    return lineHeight*unit_height


def get_xl_width(rowWidth):
    return rowWidth*unit_width


def getdata():
    global unit_width
    global unit_height
    wb = openpyxl.load_workbook('map.xlsx')

    ws = wb.worksheets[0]

    data_top, data_left, data_bottom, data_right = get_rect_from_range(
        ws.calculate_dimension())

    # print(data_top, data_left, data_bottom, data_right)

    col_width = [0]
    row_height = [0]

    for i in range(1, data_right+1):
        col_letter = openpyxl.utils.get_column_letter(i)
        col_width.append(ws.column_dimensions[col_letter].width)

    for i in range(1, data_bottom+1):
        row_height.append(ws.row_dimensions[i].height)

    unit_width = ws.column_dimensions['A'].width
    unit_height = ws.row_dimensions[1].height
    # print('UNIT:', (unit_width, unit_height))

    # print(row_height)
    # print(col_width)

    border_lines = set()
    for i in range(data_top+1, data_bottom+1):
        for j in range(data_left+1, data_right+1):
            is_merged = False
            for m in ws.merged_cells.ranges:
                if m.issuperset(openpyxl.worksheet.cell_range.CellRange(openpyxl.utils.get_column_letter(j)+str(i))):
                    # print('In merged ', m.coord, get_rect_from_range(m.coord), (i, j), get_border_lines(
                    #       ws.cell(i, j).border, get_rect_from_range(m.coord)))

                    # border_lines += get_border_lines(
                    #     ws.cell(i, j).border, get_rect_from_range(m.coord))
                    lines = get_border_lines(
                        ws.cell(i, j).border, get_rect_from_range(m.coord))
                    for l in lines:
                        border_lines.add(l)
                    is_merged = True
                    break

            if not is_merged:
                cell = ws.cell(i, j)
                coord = cell.coordinate
                # print(coord, get_border_lines(cell.border,
                #                               get_rect_from_range('%s:%s' % (coord, coord))))
                lines = get_border_lines(cell.border,
                                         get_rect_from_range('%s:%s' % (coord, coord)))
                for l in lines:
                    border_lines.add(l)

            cell = ws.cell(i, j)
            try:
                valuestr = str(cell.internal_value)
                sp = valuestr.split(',')
                rangestr = sp[0]
                directionstr = sp[1]
                rect = get_rect_from_range(rangestr)
                l = None
                left = 1
                top = 0
                right = 3
                bottom = 2
                if directionstr == 'down':
                    l = (rect[top], rect[left], rect[bottom], rect[right])
                elif directionstr == 'up':
                    l = (rect[bottom], rect[left], rect[top], rect[right])

                if l:
                    # print("DEBUG:", l, rect, rangestr)
                    border_lines.add(l)
            except Exception as e:
                pass

    # print(border_lines)

    ruller_vertical = [0, 0]
    ruller_horizontal = [0, 0]
    for i in range(2, data_bottom+1):
        height = 0
        try:
            height = float(ws.cell(i, 1).internal_value)
        except TypeError as e:
            xl_height = row_height[i]
            if xl_height == None:
                xl_height = 13.5
            height = get_real_height(xl_height)
            ws.cell(i, 1).value = str(height)

        ws.row_dimensions[i].height = get_xl_height(height)
        ruller_vertical.append(
            ruller_vertical[i-1] + height)

    for j in range(2, data_right+1):
        width = 0
        try:
            width = float(ws.cell(1, j).internal_value)
        except TypeError as e:
            xl_width = col_width[j]  # if col_width[j] is not None else 8.25
            if xl_width == None:
                xl_width = 8.25
            width = get_real_width(xl_width)
            ws.cell(1, j).value = str(width)

        col_letter = openpyxl.utils.get_column_letter(j)
        ws.column_dimensions[col_letter].width = get_xl_width(width)
        # print(width, get_xl_width(width))
        ruller_horizontal.append(
            ruller_horizontal[j-1]+width)

    # print(ruller_horizontal)
    # print(ruller_vertical)

    wb.save('map_formated.xlsx')

    # json.dump({'xs': ruller_horizontal, 'ys': ruller_vertical,
    #         'edges': [l for l in border_lines]}, open('data.json', 'w'))

    return {'xs': ruller_horizontal, 'ys': ruller_vertical,
                       'edges': [l for l in border_lines]}
