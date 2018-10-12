import openpyxl


def get_rect_from_range(range_str):
    data_range = range_str.split(':')
    left_top = openpyxl.utils.coordinate_to_tuple(data_range[0])
    right_bottom = openpyxl.utils.coordinate_to_tuple(data_range[1])
    return left_top[0]-1, left_top[1]-1, right_bottom[0], right_bottom[1]


def get_border_lines(border, rect):
    result = []

    left = 1
    top = 0
    right = 3
    bottom = 2

    if border.top.style:
        result.append((rect[top], rect[left], rect[top], rect[right]))

    if border.bottom.style:
        result.append((rect[bottom], rect[left], rect[bottom],  rect[right]))

    if border.left.style:
        result.append((rect[top], rect[left], rect[bottom],  rect[left]))

    if border.right.style:
        result.append((rect[top], rect[right], rect[bottom], rect[right]))

    if border.diagonalDown:
        result.append((rect[top], rect[left], rect[bottom], rect[right]))

    if border.diagonalUp:
        result.append((rect[bottom], rect[left], rect[top], rect[right]))

    return result


wb = openpyxl.load_workbook('m1.xlsx')

ws = wb.worksheets[0]

data_top, data_left, data_bottom, data_right = get_rect_from_range(
    ws.calculate_dimension())


print(data_top, data_left, data_bottom, data_right)

col_width = []

row_height = []

for i in range(data_left+1, data_right+1):
    col_letter = openpyxl.utils.get_column_letter(i)
    col_width.append(ws.column_dimensions[col_letter].width)

for i in range(data_top+1, data_bottom+1):
    row_height.append(ws.row_dimensions[i].height)

print(row_height)
print(col_width)

border_lines = []
for i in range(data_top+1, data_bottom+1):
    for j in range(data_left+1, data_right+1):
        is_merged = False
        for m in ws.merged_cells.ranges:
            if m.issuperset(openpyxl.worksheet.cell_range.CellRange(openpyxl.utils.get_column_letter(j)+str(i))):
                # print('In merged ', m.coord, get_rect_from_range(m.coord), (i, j), get_border_lines(
                #       ws.cell(i, j).border, get_rect_from_range(m.coord)))

                border_lines += get_border_lines(
                    ws.cell(i, j).border, get_rect_from_range(m.coord))
                is_merged = True
                break

        if not is_merged:
            cell = ws.cell(i, j)
            coord = cell.coordinate
            # print(coord, get_border_lines(cell.border,
            #                               get_rect_from_range('%s:%s' % (coord, coord))))
            border_lines += get_border_lines(cell.border,
                                             get_rect_from_range('%s:%s' % (coord, coord)))
print(border_lines)

